import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Color
import colorsys
import math

def generate_distinct_light_colors(num_colors):
    """
    Generate a set of visually distinct light colors
    Uses a more sophisticated color distribution method
    """
    light_colors = []

    # Define a base set of light color palettes
    color_palettes = [
        # Pastel blues
        [(210, 0.2, 0.9), (220, 0.2, 0.85), (230, 0.2, 0.95)],
        # Pastel greens
        [(120, 0.2, 0.9), (130, 0.2, 0.85), (110, 0.2, 0.95)],
        # Pastel purples
        [(270, 0.2, 0.9), (280, 0.2, 0.85), (260, 0.2, 0.95)],
        # Pastel pinks
        [(330, 0.2, 0.9), (340, 0.2, 0.85), (320, 0.2, 0.95)],
        # Pastel yellows
        [(50, 0.2, 0.9), (60, 0.2, 0.85), (40, 0.2, 0.95)],
        # Pastel oranges
        [(20, 0.2, 0.9), (30, 0.2, 0.85), (10, 0.2, 0.95)]
    ]

    # Cycle through palettes if more colors are needed
    for i in range(num_colors):
        # Select palette and color within the palette
        palette = color_palettes[i % len(color_palettes)]
        color_variant = palette[i % len(palette)]

        # Convert HSV to RGB
        hue = color_variant[0] / 360  # Normalize hue
        saturation = color_variant[1]
        value = color_variant[2]

        rgb = colorsys.hsv_to_rgb(hue, saturation, value)

        # Convert RGB (0-1) to hex
        hex_color = ''.join([f'{int(x*255):02x}' for x in rgb])
        light_colors.append(hex_color)

    return light_colors

def add_repository_borders_and_distinct_colors(input_csv, output_excel):
    # Read the CSV file
    df = pd.read_csv(input_csv)

    # Sort the dataframe by Repository Name to group repositories together
    df_sorted = df.sort_values('Repository Name')

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers
    headers = df_sorted.columns.tolist()
    for col, header in enumerate(headers, 1):
        header_cell = ws.cell(row=1, column=col, value=header)
        header_cell.font = openpyxl.styles.Font(bold=True)

    # Define border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Generate distinct light colors for repositories
    unique_repos = df_sorted['Repository Name'].unique()
    repo_colors = generate_distinct_light_colors(len(unique_repos))
    repo_color_map = dict(zip(unique_repos, repo_colors))

    # Track the current repository and row to write
    current_repo = None
    write_row = 2

    # Iterate through sorted dataframe
    for index, row in df_sorted.iterrows():
        # Check if repository has changed
        if row['Repository Name'] != current_repo:
            # If not the first repository, add an empty row
            if current_repo is not None:
                write_row += 1

            # Update current repository
            current_repo = row['Repository Name']

        # Get the distinct light color for this repository
        fill_color = repo_color_map[current_repo]
        light_fill = PatternFill(start_color=fill_color,
                                 end_color=fill_color,
                                 fill_type='solid')

        # Write data to the worksheet
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=write_row, column=col, value=value)
            cell.border = thin_border
            cell.fill = light_fill

        # Move to next row
        write_row += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(output_excel)
    print(f"Processed file saved as {output_excel}")

# Example usage
input_csv = 'reposcsv1.csv'
output_excel = 'repositories_with_distinct_colors.xlsx'
add_repository_borders_and_distinct_colors(input_csv, output_excel)