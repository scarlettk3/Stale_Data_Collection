import os
import time
import json
import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
from datetime import datetime, timezone
from tqdm import tqdm  # For progress bars
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.styles import Font, PatternFill, Border, Side

# Configuration
GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN') or 'place_you_github_token_here'
ORGANIZATION = 'place_you_organisation_name'
INPUT_CSV = 'github_repo_analysis_with_stale.csv'  # Your CSV file with repository information
OUTPUT_EXCEL = 'github_stale_information.xlsx'  # Output Excel file with multiple sheets
CHECKPOINT_FILE = 'stale_branch_detailed_checkpoint.json'  # For saving progress

# Column names in your CSV file
REPO_COLUMN = 'repository_name'  # Based on your CSV
BRANCH_COUNT_COLUMN = 'number_of_branches'  # Column with total branch count
STALE_BRANCH_COUNT_COLUMN = 'Stale_Branches'  # Column with stale branch count

# Headers for GitHub API requests
headers = {
    'Authorization': f'token {GITHUB_TOKEN}',
    'Accept': 'application/vnd.github.v3+json'
}

# Rate limit configuration
API_CALL_DELAY = 0.2  # 200ms delay between API calls
BATCH_SIZE = 5  # Process fewer repos per batch since they're larger
BATCH_BREAK = 120  # Longer break between batches
CONNECTION_TIMEOUT = 45  # Increased timeout for API requests in seconds
MAX_RETRIES = 5  # Increased maximum number of retries for API calls

def create_session():
    """Create a requests session with retry configuration."""
    session = requests.Session()
    retries = Retry(
        total=7,  # Increased retry limit
        backoff_factor=1.5,  # Increased backoff factor
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"]
    )
    session.mount("https://", HTTPAdapter(max_retries=retries))
    return session

def check_rate_limit(session, wait_if_needed=True):
    """Check GitHub API rate limit status and wait if needed."""
    url = 'https://api.github.com/rate_limit'
    try:
        response = session.get(url, headers=headers, timeout=CONNECTION_TIMEOUT)

        if response.status_code != 200:
            print(f"Error checking rate limit: {response.status_code}")
            return False

        data = response.json()
        remaining = data['resources']['core']['remaining']
        reset_time = data['resources']['core']['reset']

        print(f"API calls remaining: {remaining}")

        if remaining < 200 and wait_if_needed:  # Increased threshold
            reset_timestamp = int(reset_time)
            current_timestamp = int(time.time())
            sleep_time = reset_timestamp - current_timestamp + 15  # Added 15 seconds buffer

            if sleep_time > 0:
                print(f"\nRate limit low ({remaining} remaining). Waiting {sleep_time} seconds until reset...")

                # Progress indicator while waiting
                for i in tqdm(range(sleep_time), desc="Waiting for rate limit reset"):
                    time.sleep(1)

                print("Continuing with API requests...")
                return True

        return True
    except Exception as e:
        print(f"Error in check_rate_limit: {str(e)}")
        return False

def safe_api_call(session, url, retry_count=0):
    """Make an API call with error handling and retries."""
    try:
        response = session.get(url, headers=headers, timeout=CONNECTION_TIMEOUT)
        time.sleep(API_CALL_DELAY)  # Rate limiting
        return response
    except (requests.exceptions.SSLError, requests.exceptions.ConnectionError,
            requests.exceptions.Timeout, requests.exceptions.RequestException) as e:
        if retry_count < MAX_RETRIES:
            wait_time = 5 * (2 ** retry_count)  # Exponential backoff
            print(f"Connection error, retrying in {wait_time}s ({retry_count+1}/{MAX_RETRIES}): {str(e)}")
            time.sleep(wait_time)
            return safe_api_call(session, url, retry_count + 1)
        else:
            print(f"Failed after {MAX_RETRIES} retries: {str(e)}")
            return None

def load_checkpoint():
    """Load checkpoint data from file if it exists."""
    if os.path.exists(CHECKPOINT_FILE):
        try:
            with open(CHECKPOINT_FILE, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading checkpoint file: {str(e)}")
    return {}

def save_checkpoint(checkpoint_data):
    """Save checkpoint data to file."""
    try:
        with open(CHECKPOINT_FILE, 'w') as f:
            json.dump(checkpoint_data, f)
        print(f"Checkpoint saved to {CHECKPOINT_FILE}")
    except Exception as e:
        print(f"Error saving checkpoint file: {str(e)}")

def find_last_merged_branch(session, repo_full_name, branch_name):
    """Find the branch that this branch was last merged to, based on pull request data."""
    try:
        # Look for pull requests where this branch was merged
        # First, try to find PRs that mention the branch name
        search_query = f"repo:{repo_full_name} head:{branch_name} is:pr is:merged"
        search_url = f"https://api.github.com/search/issues?q={search_query}"

        response = session.get(search_url, headers=headers, timeout=CONNECTION_TIMEOUT)
        time.sleep(API_CALL_DELAY)

        if response.status_code == 200:
            data = response.json()

            if data.get('total_count', 0) > 0:
                # Get the first (most recent) PR
                pr = data['items'][0]
                pr_number = pr['number']

                # Get the PR details to find the base branch (where it was merged to)
                pr_url = f"https://api.github.com/repos/{repo_full_name}/pulls/{pr_number}"
                pr_response = session.get(pr_url, headers=headers, timeout=CONNECTION_TIMEOUT)
                time.sleep(API_CALL_DELAY)

                if pr_response.status_code == 200:
                    pr_data = pr_response.json()
                    base_branch = pr_data.get('base', {}).get('ref')
                    if base_branch:
                        return base_branch

        # If the above approach didn't work, try to find commits that merged this branch
        # This is a fallback to your original approach but with improved regex patterns
        search_query = f"repo:{repo_full_name} merge {branch_name} type:commit"
        search_url = f"https://api.github.com/search/commits?q={search_query}"

        search_headers = headers.copy()
        search_headers['Accept'] = 'application/vnd.github.cloak-preview+json'

        response = session.get(search_url, headers=search_headers, timeout=CONNECTION_TIMEOUT)
        time.sleep(API_CALL_DELAY)

        if response.status_code == 200:
            data = response.json()

            if data.get('total_count', 0) > 0:
                for commit in data['items'][:3]:  # Check a few commits
                    commit_message = commit['commit']['message']

                    # Try various patterns used in merge commits
                    patterns = [
                        r"Merge (?:pull request|PR) #\d+ .*?into ([^\s]+)",  # PR merges
                        r"Merge branch '?([^']+)'? into ([^\s']+)",  # Branch merges
                        r"Merge '?([^']+)'? into ([^\s']+)",  # Other merge format
                        r"merged \d+ commit\(s\) into ([^\s]+) from",  # GitHub format
                        r"from .* into ([^\s]+)"  # Another GitHub format
                    ]

                    for pattern in patterns:
                        match = re.search(pattern, commit_message)
                        if match:
                            # Some patterns have the target branch in group 1, others in group 2
                            target_branch = match.group(1) if len(match.groups()) == 1 else match.group(2)
                            return target_branch

        # If we still haven't found anything, we'll check the repo's default branch
        # as a last resort (common branches that stale branches might have been merged to)
        repo_url = f"https://api.github.com/repos/{repo_full_name}"
        repo_response = session.get(repo_url, headers=headers, timeout=CONNECTION_TIMEOUT)

        if repo_response.status_code == 200:
            repo_data = repo_response.json()
            default_branch = repo_data.get('default_branch')

            # Look for any merge to the default branch that might involve this branch
            search_query = f"repo:{repo_full_name} {branch_name} {default_branch} type:commit"
            search_url = f"https://api.github.com/search/commits?q={search_query}"

            response = session.get(search_url, headers=search_headers, timeout=CONNECTION_TIMEOUT)

            if response.status_code == 200:
                data = response.json()
                if data.get('total_count', 0) > 0:
                    # If we found commits mentioning both branches, it's likely a merge happened
                    return default_branch

        # If we couldn't find merge information, return Unknown
        return "Unknown"
    except Exception as e:
        print(f"Error finding merge history for {branch_name}: {str(e)}")
        return "Error"

def get_stale_branches_info(session, repo_full_name, repo_stale_count, checkpoint_data):
    """Get detailed information about stale branches using REST API with checkpointing."""
    if not check_rate_limit(session):
        return None, "Rate Limit Error"

    # Check if we have checkpoint data for this repo
    repo_checkpoint = checkpoint_data.get(repo_full_name, {})
    processed_branches = repo_checkpoint.get('processed_branches', [])
    stale_branches_info = repo_checkpoint.get('stale_branches_info', [])

    # If we already found enough stale branches to match the expected count, we can skip further processing
    if stale_branches_info and len(stale_branches_info) >= repo_stale_count:
        print(f"Already found {len(stale_branches_info)} stale branches for {repo_full_name}, which matches or exceeds the expected {repo_stale_count}")
        return stale_branches_info, "Completed"

    # Get the default branch to exclude it from stale count
    repo_url = f'https://api.github.com/repos/{repo_full_name}'
    repo_response = safe_api_call(session, repo_url)

    if repo_response is None or repo_response.status_code != 200:
        print(f"Error fetching repo info for {repo_full_name}: {getattr(repo_response, 'status_code', 'N/A')}")
        default_branch = 'main'  # Fallback to a common default branch name
    else:
        repo_data = repo_response.json()
        default_branch = repo_data.get('default_branch', 'main')

    # Get all branches - we'll paginate through ALL branches
    all_branches = []
    page = 1
    pages_processed = repo_checkpoint.get('pages_processed', 0)

    if pages_processed > 0:
        print(f"Resuming from page {pages_processed + 1} for {repo_full_name}")
        page = pages_processed + 1
    else:
        print(f"Starting to fetch branches for {repo_full_name}")

    while True:
        branches_url = f'https://api.github.com/repos/{repo_full_name}/branches?per_page=100&page={page}'
        branches_response = safe_api_call(session, branches_url)

        if branches_response is None:
            return stale_branches_info, "Connection Error"
        elif branches_response.status_code == 404:
            print(f"Repository {repo_full_name} not found")
            return stale_branches_info, "Repo Not Found"
        elif branches_response.status_code != 200:
            print(f"Error fetching branches for {repo_full_name}: {branches_response.status_code}")
            return stale_branches_info, "Error"

        branches = branches_response.json()
        if not branches:
            break  # No more branches

        print(f"Fetched page {page} with {len(branches)} branches for {repo_full_name}")
        all_branches.extend(branches)

        # Update page counter in checkpoint
        repo_checkpoint['pages_processed'] = page
        checkpoint_data[repo_full_name] = repo_checkpoint
        save_checkpoint(checkpoint_data)

        page += 1

        # Check rate limit after each page of branches
        check_rate_limit(session)

        # If we've found enough stale branches already, we can stop fetching more pages
        if len(stale_branches_info) >= repo_stale_count:
            print(f"Found enough stale branches ({len(stale_branches_info)}) to match the expected count ({repo_stale_count})")
            break

    # Current time in seconds since epoch
    current_time = time.time()

    # 90 days in seconds
    stale_threshold = 90 * 24 * 60 * 60

    print(f"Fetched a total of {len(all_branches)} branches for {repo_full_name}")

    # Create a list of branches to process (excluding those already processed)
    branches_to_process = [branch for branch in all_branches if branch['name'] not in processed_branches]
    print(f"Processing {len(branches_to_process)} remaining branches (already processed {len(processed_branches)})")

    # Process branches in smaller chunks for large repositories
    chunk_size = min(50, len(branches_to_process))
    chunks_total = (len(branches_to_process) + chunk_size - 1) // chunk_size

    for chunk_idx in range(chunks_total):
        chunk_start = chunk_idx * chunk_size
        chunk_end = min(chunk_start + chunk_size, len(branches_to_process))
        chunk = branches_to_process[chunk_start:chunk_end]

        print(f"Processing chunk {chunk_idx + 1}/{chunks_total} ({chunk_end - chunk_start} branches)")

        # Use tqdm for a progress bar within this chunk
        for branch in tqdm(chunk, desc=f"Checking branches {chunk_start+1}-{chunk_end}/{len(branches_to_process)}"):
            # Skip default branch
            if branch['name'] == default_branch:
                processed_branches.append(branch['name'])
                continue

            # Get the latest commit on this branch
            commit_url = f'https://api.github.com/repos/{repo_full_name}/commits/{branch["commit"]["sha"]}'
            commit_response = safe_api_call(session, commit_url)

            if commit_response is None:
                continue  # Skip this branch if we couldn't get the commit info

            if commit_response.status_code != 200:
                print(f"Error fetching commit for branch {branch['name']}: {commit_response.status_code}")
                continue

            try:
                commit_data = commit_response.json()

                # Get commit date
                commit_date_str = commit_data['commit']['committer']['date']
                commit_date = datetime.strptime(commit_date_str, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
                commit_timestamp = commit_date.timestamp()
                formatted_date = commit_date.strftime("%Y-%m-%d %H:%M:%S")

                # Check if older than 90 days
                if (current_time - commit_timestamp) > stale_threshold:
                    # Find where this branch was last merged to
                    print(f"Finding merge history for stale branch: {branch['name']}")
                    last_merged_to = find_last_merged_branch(session, repo_full_name, branch['name'])

                    # Add this stale branch info to our list
                    stale_branches_info.append({
                        'branch_name': branch['name'],
                        'last_commit_date': formatted_date,
                        'last_merged_to': last_merged_to
                    })

                    # Update checkpoint after finding each stale branch
                    repo_checkpoint['stale_branches_info'] = stale_branches_info
                    checkpoint_data[repo_full_name] = repo_checkpoint
                    save_checkpoint(checkpoint_data)

                    # If we've found enough stale branches, we can stop
                    if len(stale_branches_info) >= repo_stale_count:
                        print(f"Found enough stale branches ({len(stale_branches_info)}) to match the expected count ({repo_stale_count})")
                        return stale_branches_info, "Completed"
            except Exception as e:
                print(f"Error processing branch {branch['name']}: {str(e)}")

            # Mark this branch as processed
            processed_branches.append(branch['name'])

            # Update checkpoint after each branch
            repo_checkpoint['processed_branches'] = processed_branches
            checkpoint_data[repo_full_name] = repo_checkpoint
            save_checkpoint(checkpoint_data)

        # Take a break between chunks
        if chunk_idx < chunks_total - 1:
            print("Taking a short break between chunks...")
            time.sleep(5)
            check_rate_limit(session)

    print(f"Found {len(stale_branches_info)} stale branches in {repo_full_name}")
    return stale_branches_info, "Completed"

def check_excel_sheet_exists(output_file, repo_name):
    """Check if a sheet for this repository exists in the Excel file and has data."""
    if not os.path.exists(output_file):
        return False

    try:
        # Make a safe sheet name (same as in create_or_update_excel)
        sheet_name = repo_name[:31]
        sheet_name = re.sub(r'[\[\]\:\*\?\/\\]', '_', sheet_name)

        # Load the workbook
        workbook = openpyxl.load_workbook(output_file, read_only=True)

        # Check if sheet exists
        if sheet_name not in workbook.sheetnames:
            return False

        # Get the sheet
        sheet = workbook[sheet_name]

        # Check if it has data (more than just headers)
        return sheet.max_row > 5  # We expect at least one row of data beyond headers

    except Exception as e:
        print(f"Error checking Excel sheet for {repo_name}: {str(e)}")
        return False

def create_or_update_excel(output_file, repo_name, branch_count, stale_branch_count, stale_branches_info):
    """Create or update an Excel file with a sheet for this repository."""
    try:
        # Check if the Excel file already exists
        if os.path.exists(output_file):
            # Load existing workbook
            workbook = openpyxl.load_workbook(output_file)
        else:
            # Create a new workbook
            workbook = openpyxl.Workbook()
            # Remove the default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

        # Create a safe sheet name (Excel limits sheet names to 31 chars and certain characters are not allowed)
        sheet_name = repo_name[:31]  # Truncate to 31 chars
        sheet_name = re.sub(r'[\[\]\:\*\?\/\\]', '_', sheet_name)  # Replace invalid chars

        # Check if sheet already exists
        if sheet_name in workbook.sheetnames:
            # Use the existing sheet
            sheet = workbook[sheet_name]
            sheet.delete_rows(1, sheet.max_row)  # Clear existing content
        else:
            # Create a new sheet
            sheet = workbook.create_sheet(title=sheet_name)

        # Apply some formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add repository info at the top using the values from the CSV
        sheet['A1'] = 'Repository Name'
        sheet['B1'] = repo_name
        sheet['A1'].font = header_font
        sheet['A2'] = 'Total Branches'
        sheet['B2'] = branch_count
        sheet['A2'].font = header_font
        sheet['A3'] = 'Stale Branches Count'
        sheet['B3'] = stale_branch_count
        sheet['A3'].font = header_font
        # Add this to your create_or_update_excel function after line 377 (after adding repository info)
        # Create a hyperlink back to the master sheet
        sheet['D1'] = "Back to Master Sheet"
        sheet['D1'].hyperlink = "#'Master Sheet'!A1"
        sheet['D1'].style = "Hyperlink"
        # Add headers for stale branches table
        table_headers = ['Stale Branch Name', 'Last Commit Date', 'Last Merged To']
        for i, header in enumerate(table_headers, 1):
            cell = sheet.cell(row=5, column=i)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Add stale branch data
        for i, branch_info in enumerate(stale_branches_info, 6):
            sheet.cell(row=i, column=1).value = branch_info['branch_name']
            sheet.cell(row=i, column=2).value = branch_info['last_commit_date']
            sheet.cell(row=i, column=3).value = branch_info['last_merged_to']

            # Apply border to cells
            for col in range(1, 4):
                sheet.cell(row=i, column=col).border = thin_border

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length > 0 else 10
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Limit to 50 for very long values

        # Save the workbook
        workbook.save(output_file)
        print(f"Updated sheet '{sheet_name}' in {output_file}")
        return True
    except Exception as e:
        print(f"Error creating/updating Excel file for {repo_name}: {str(e)}")
        return False
def create_master_sheet(output_file, INPUT_CSV):
    """Create a master sheet with links to all repository tabs, reading data from CSV."""
    try:
        # Load the CSV data
        df = pd.read_csv(INPUT_CSV)
        print(f"Loaded {len(df)} repositories from {INPUT_CSV} for master sheet")

        # Load the workbook
        workbook = openpyxl.load_workbook(output_file)

        # Check if master sheet already exists, remove it if it does
        if 'Master Sheet' in workbook.sheetnames:
            del workbook['Master Sheet']

        # Create a new master sheet and position it as the first sheet
        master_sheet = workbook.create_sheet('Master Sheet', 0)

        # Define headers for the master table
        headers = ['Repository Name', 'Total Branches', 'Stale Branches', 'Link to Details']

        # Apply formatting to headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add headers to the master sheet
        for col, header in enumerate(headers, 1):
            cell = master_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Initialize row counter for master sheet
        master_row = 2

        # Add each repository from the CSV to the master sheet
        for index, row in df.iterrows():
            repo_name = row[REPO_COLUMN]
            branch_count = int(row[BRANCH_COUNT_COLUMN])
            stale_branch_count = int(row[STALE_BRANCH_COUNT_COLUMN])

            # Add this repository to the master sheet
            master_sheet.cell(row=master_row, column=1).value = repo_name
            master_sheet.cell(row=master_row, column=2).value = branch_count
            master_sheet.cell(row=master_row, column=3).value = stale_branch_count

            # Create a safe sheet name (same as in create_or_update_excel)
            sheet_name = repo_name[:31]  # Truncate to 31 chars
            sheet_name = re.sub(r'[\[\]\:\*\?\/\\]', '_', sheet_name)  # Replace invalid chars

            # Add a hyperlink if the sheet exists
            if sheet_name in workbook.sheetnames:
                link_cell = master_sheet.cell(row=master_row, column=4)
                link_cell.value = "Go to details"
                link_cell.hyperlink = f"#{sheet_name}!A1"
                link_cell.style = "Hyperlink"  # Use Excel's built-in hyperlink style
            else:
                master_sheet.cell(row=master_row, column=4).value = "No details available"

            # Apply borders to the cells
            for col in range(1, 5):
                master_sheet.cell(row=master_row, column=col).border = thin_border

            master_row += 1

        # Add a count of repositories
        master_sheet.cell(row=master_row + 1, column=1).value = f"Total Repositories: {master_row - 2}"
        master_sheet.cell(row=master_row + 1, column=1).font = Font(bold=True)

        # Auto-adjust column widths for the master sheet
        for column in master_sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length > 0 else 10
            master_sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

        # Add filters to the header row to enable sorting and filtering
        master_sheet.auto_filter.ref = f"A1:D{master_row-1}"

        # Save the workbook
        workbook.save(output_file)
        print(f"Created master sheet with {master_row-2} repositories from CSV data.")
        return True
    except Exception as e:
        print(f"Error creating master sheet: {str(e)}")
        return False

def main():
    print(f"Starting detailed stale branch analysis for repositories in {ORGANIZATION}")
    print(f"Using input file: {INPUT_CSV}")
    print(f"Results will be saved to: {OUTPUT_EXCEL}")
    print(f"Using checkpoint file: {CHECKPOINT_FILE}")

    # Create a session with retry configuration
    session = create_session()

    # Load checkpoint data
    checkpoint_data = load_checkpoint()

    # Check rate limit before starting
    check_rate_limit(session, wait_if_needed=True)

    # Load existing CSV
    try:
        df = pd.read_csv(INPUT_CSV)
        print(f"Loaded {len(df)} repositories from {INPUT_CSV}")
    except Exception as e:
        print(f"Error loading CSV: {str(e)}")
        return

    # Verify the required columns exist
    required_columns = [REPO_COLUMN, BRANCH_COUNT_COLUMN, STALE_BRANCH_COUNT_COLUMN]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        print(f"Error: The following required columns are missing from the CSV: {', '.join(missing_columns)}")
        print(f"Available columns are: {', '.join(df.columns)}")
        return

    # Find repositories that need processing or need to be saved to Excel
    repos_to_process = []
    for index, row in df.iterrows():
        repo_name = row[REPO_COLUMN]
        branch_count = int(row[BRANCH_COUNT_COLUMN])
        stale_branch_count = int(row[STALE_BRANCH_COUNT_COLUMN])

        # Skip repositories with no stale branches
        if stale_branch_count == 0:
            print(f"Skipping {repo_name} as it has no stale branches")
            # Still create an Excel sheet for it with zero stale branches
            if not check_excel_sheet_exists(OUTPUT_EXCEL, repo_name):
                create_or_update_excel(OUTPUT_EXCEL, repo_name, branch_count, stale_branch_count, [])
            continue

        # Check if this repo is already completely processed
        repo_full_name = f"{ORGANIZATION}/{repo_name}"
        repo_checkpoint = checkpoint_data.get(repo_full_name, {})

        stale_branches_info = repo_checkpoint.get('stale_branches_info', [])

        if (stale_branches_info and len(stale_branches_info) >= stale_branch_count):
            print(f"Repository {repo_name} already processed in checkpoint.")

            # Check if the data is also in Excel
            if check_excel_sheet_exists(OUTPUT_EXCEL, repo_name):
                print(f"Repository {repo_name} already has Excel sheet. Skipping.")
                continue
            else:
                print(f"Repository {repo_name} processed but Excel sheet missing. Creating sheet.")
                create_or_update_excel(OUTPUT_EXCEL, repo_name, branch_count, stale_branch_count, stale_branches_info)
                continue

        # Otherwise, add to processing list
        repos_to_process.append((index, repo_name, branch_count, stale_branch_count))

    print(f"Need to process {len(repos_to_process)} repositories")

    # Process repositories in batches
    for batch_idx, batch in enumerate(range(0, len(repos_to_process), BATCH_SIZE)):
        batch_repos = repos_to_process[batch:batch + BATCH_SIZE]

        print(f"\n--- Processing Batch {batch_idx + 1} ({len(batch_repos)} repositories) ---")

        # Process each repository in the batch
        for i, (index, repo_name, branch_count, stale_branch_count) in enumerate(batch_repos):
            repo_full_name = f"{ORGANIZATION}/{repo_name}"
            print(f"\nProcessing {i+1}/{len(batch_repos)} in batch: {repo_name}")
            print(f"Overall progress: {batch + i + 1}/{len(repos_to_process)}")
            print(f"Expected branches: {branch_count}, Expected stale branches: {stale_branch_count}")

            # Get stale branches information
            stale_branches_info, status = get_stale_branches_info(session, repo_full_name, stale_branch_count, checkpoint_data)

            if stale_branches_info is not None:
                # Create or update the Excel sheet for this repository
                success = create_or_update_excel(OUTPUT_EXCEL, repo_name, branch_count, stale_branch_count, stale_branches_info)
                if not success:
                    print(f"WARNING: Failed to create/update Excel sheet for {repo_name}.")

                    # Add a "needs_excel_update" flag to the checkpoint
                    repo_checkpoint = checkpoint_data.get(repo_full_name, {})
                    repo_checkpoint['needs_excel_update'] = True
                    checkpoint_data[repo_full_name] = repo_checkpoint
                    save_checkpoint(checkpoint_data)
            else:
                print(f"Error processing {repo_name}. Status: {status}")

            # Take a break between repos within a batch
            if i < len(batch_repos) - 1:
                print("Taking a short break between repositories...")
                time.sleep(10)
                check_rate_limit(session)

        # After each batch, take a break and check rate limits
        if batch + BATCH_SIZE < len(repos_to_process):
            print(f"\nCompleted batch {batch_idx + 1}. Taking a {BATCH_BREAK} second break...")

            # Wait with progress bar
            for _ in tqdm(range(BATCH_BREAK), desc="Batch break"):
                time.sleep(1)

            # Check rate limit before starting next batch
            check_rate_limit(session, wait_if_needed=True)

    # Check for any repos that were processed but failed to save to Excel
    print("\nChecking for repositories that need Excel updates...")
    excel_updates_needed = []

    for repo_full_name, repo_data in checkpoint_data.items():
        if repo_data.get('needs_excel_update', False) or (
                repo_data.get('stale_branches_info') and
                not check_excel_sheet_exists(OUTPUT_EXCEL, repo_full_name.split('/')[-1])
        ):
            excel_updates_needed.append(repo_full_name)

    if excel_updates_needed:
        print(f"Found {len(excel_updates_needed)} repositories that need Excel updates.")

        for repo_full_name in excel_updates_needed:
            repo_name = repo_full_name.split('/')[-1]
            print(f"Updating Excel for {repo_name}...")

            # Find the repo in the DataFrame to get branch counts
            repo_row = df[df[REPO_COLUMN] == repo_name]
            if repo_row.empty:
                print(f"Warning: Could not find {repo_name} in CSV file.")
                continue

            branch_count = int(repo_row[BRANCH_COUNT_COLUMN].iloc[0])
            stale_branch_count = int(repo_row[STALE_BRANCH_COUNT_COLUMN].iloc[0])

            # Get the stale branches info from checkpoint
            repo_data = checkpoint_data[repo_full_name]
            stale_branches_info = repo_data.get('stale_branches_info', [])

            # Create or update the Excel sheet
            success = create_or_update_excel(OUTPUT_EXCEL, repo_name, branch_count, stale_branch_count, stale_branches_info)

            if success:
                # Update checkpoint to remove the flag
                repo_data['needs_excel_update'] = False
                checkpoint_data[repo_full_name] = repo_data
                save_checkpoint(checkpoint_data)
    else:
        print("All repositories have been properly saved to Excel.")

    # Print summary, After processing all repositories and updating Excel sheets
    print("\n=== Detailed Stale Branch Analysis Complete ===")
    print(f"Results saved to {OUTPUT_EXCEL}")



    # Create the master sheet with links, passing in the CSV file
    print("\nCreating master sheet with links to repository details...")
    create_master_sheet(OUTPUT_EXCEL, INPUT_CSV)

    print("Analysis complete. You can now access all repositories from the Master Sheet.")

if __name__ == "__main__":
    main()



